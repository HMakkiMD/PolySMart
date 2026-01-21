import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook('analysis.xlsx')

# Select the 'p3hb_bulk' sheet
sheet = workbook['p3hb_bulk']

# Iterate through rows and copy values from column 3 to column 4
for row in sheet.iter_rows(min_row=1, min_col=3, max_col=3):
    for cell in row:
        # Copy value from column 3 to column 4
        sheet.cell(row=cell.row, column=4, value=cell.value)
        # Replace value in column 3 with zero
        sheet.cell(row=cell.row, column=3, value=0)

# Save the changes
workbook.save('analysis.xlsx')
