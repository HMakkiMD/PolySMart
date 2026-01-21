from openpyxl import load_workbook

# Load the workbook
file_name = "analysis.xlsx"
wb = load_workbook(file_name)

# Sheets to delete
sheets_to_delete = ["Sheet1", "p3hb_bulk"]

for sheet_name in sheets_to_delete:
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        print(f"Deleted sheet: {sheet_name}")
    else:
        print(f"Sheet not found: {sheet_name}")

# Sheets to rename (old_name : new_name)
sheets_to_rename = {
    "interface_and_pe": "reacted clusters",
    "P3HB_PE together": "gMBCPs composition"
}

for old_name, new_name in sheets_to_rename.items():
    if old_name in wb.sheetnames:
        wb[old_name].title = new_name
        print(f"Renamed sheet: {old_name} -> {new_name}")
    else:
        print(f"Sheet not found for renaming: {old_name}")

# Save the workbook
wb.save(file_name)
print(f"Workbook saved: {file_name}")



import os

# List of files to delete
files_to_delete = [
    "CRE-cluster.xlsx",
    "PBAT-cluster.xlsx",
    "PLA-cluster.xlsx",
    "p3hb_bulk.txt",
    "interface_and_pe.txt"
]

for file_name in files_to_delete:
    if os.path.exists(file_name):
        try:
            os.remove(file_name)
            print(f"Deleted: {file_name}")
        except Exception as e:
            print(f"Error deleting {file_name}: {e}")
    else:
        print(f"File not found: {file_name}")
